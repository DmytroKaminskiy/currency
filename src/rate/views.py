import csv
from datetime import datetime
from urllib.parse import urlencode
from django.db.models import F

from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import DeleteView, ListView, TemplateView, UpdateView, View

from openpyxl import Workbook

from django_filters.views import FilterView

from rate.models import Rate, Track
from rate.selectors import get_latest_rates
from rate.utils import display
from rate.filters import RateFilter


class RateList(FilterView):
    queryset = Rate.objects.all()
    template_name = 'rate/rate_list.html'
    paginate_by = 25
    filterset_class = RateFilter

    def get(self, request, *args, **kwargs):
        print('GET')
        # track, created = Track.objects.get_or_create()
        # from time import sleep
        # print(track.counter)
        # sleep(30)
        # track.counter = track.counter + 1
        # track.save()
        Track.objects.all().update(counter=F('counter') + 1)

        return super().get(request, *args, **kwargs)


    # def get_context_data(self, *args, **kwargs):
    #     context = super().get_context_data(*args, **kwargs)
    #     get = dict(tuple(self.request.GET.items()))
    #     get.pop('page', None)
    #     context['request_GET'] = urlencode(get)
    #     return context


class LatestRatesView(TemplateView):
    template_name = 'latest-rates.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_list'] = get_latest_rates()
        return context


class RateDownloadCSV(View):
    HEADERS = (
        'id',
        'created',
        'source',
        'amount',
        'type',
    )
    queryset = Rate.objects.all().iterator()

    def get(self, request):
        # Create the HttpResponse object with the appropriate CSV header.
        response = self.get_response()

        writer = csv.writer(response)  # response.write
        writer.writerow(self.__class__.HEADERS)

        for rate in self.queryset:
            # row
            values = []
            for attr in self.__class__.HEADERS:
                values.append(display(rate, attr))

            writer.writerow(values)

        return response

    def get_response(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rate.csv"'
        return response


class RateDownloadXLSX(View):
    columns = [
        'id',
        'created',
        'source',
        'amount',
        'type',
    ]
    queryset = Rate.objects.all().iterator()

    def get(self, request):
        """
        Downloads all movies as Excel file with a single worksheet
        """
        # movie_queryset = Movie.objects.all()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        today = datetime.now().strftime("%Y-%m-%d")
        response['Content-Disposition'] = f'attachment; filename={today}-rates.xlsx'
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Movies'

        # Define the titles for columns
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for rate in self.queryset:
            row_num += 1

            # row
            row = []
            for attr in self.__class__.columns:
                row.append(display(rate, attr))

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


class UpdateRate(UserPassesTestMixin, UpdateView):
    template_name = 'rate-update.html'
    queryset = Rate.objects.all()
    fields = ('amount', 'source', 'currency_type', 'type')
    success_url = reverse_lazy('rate:list')

    def test_func(self):
        return self.request.user.is_authenticated and\
            self.request.user.is_superuser


class DeleteRate(UserPassesTestMixin, DeleteView):
    queryset = Rate.objects.all()
    success_url = reverse_lazy('rate:list')

    def get(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)

    def test_func(self):
        return self.request.user.is_authenticated and\
            self.request.user.is_superuser
